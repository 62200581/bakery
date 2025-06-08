import pandas as pd
from openpyxl import load_workbook
import os

def add_data_to_excel():
    """Add sample data to an Excel file"""
    excel_file = 'bakery_data.xlsx'
    
    # Check if file exists
    file_exists = os.path.isfile(excel_file)
    
    if file_exists:
        print(f"Adding data to existing file: {excel_file}")
        # Load existing workbook
        book = load_workbook(excel_file)
        
        # Check if sheet exists, otherwise create new one
        sheet_name = 'new_products'
        if sheet_name in book.sheetnames:
            print(f"Sheet '{sheet_name}' already exists. Appending data.")
            # Read existing data
            existing_df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            # Create new data
            new_data = {
                'ProductID': ['P1001', 'P1002', 'P1003'],
                'ProductName': ['Chocolate Croissant', 'Strawberry Tart', 'Focaccia Bread'],
                'Category': ['Pastry', 'Dessert', 'Bread'],
                'Price': [3.99, 4.50, 5.99],
                'Stock': [25, 15, 20]
            }
            new_df = pd.DataFrame(new_data)
            
            # Combine data (append new rows)
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)
            
            # Write combined data back to the same sheet
            with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
            print(f"Added {len(new_df)} new rows to '{sheet_name}' sheet")
        else:
            print(f"Creating new sheet: '{sheet_name}'")
            # Create new data
            new_data = {
                'ProductID': ['P1001', 'P1002', 'P1003'],
                'ProductName': ['Chocolate Croissant', 'Strawberry Tart', 'Focaccia Bread'],
                'Category': ['Pastry', 'Dessert', 'Bread'],
                'Price': [3.99, 4.50, 5.99],
                'Stock': [25, 15, 20]
            }
            new_df = pd.DataFrame(new_data)
            
            # Write to new sheet
            with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
                new_df.to_excel(writer, sheet_name=sheet_name, index=False)
                
            print(f"Added {len(new_df)} rows to new sheet '{sheet_name}'")
    else:
        print(f"Creating new Excel file: {excel_file}")
        # Create new data
        new_data = {
            'ProductID': ['P1001', 'P1002', 'P1003'],
            'ProductName': ['Chocolate Croissant', 'Strawberry Tart', 'Focaccia Bread'],
            'Category': ['Pastry', 'Dessert', 'Bread'],
            'Price': [3.99, 4.50, 5.99],
            'Stock': [25, 15, 20]
        }
        new_df = pd.DataFrame(new_data)
        
        # Write to new file
        new_df.to_excel(excel_file, sheet_name='new_products', index=False)
        print(f"Created new file with {len(new_df)} rows in 'new_products' sheet")
    
    print("Excel operation completed successfully")

def create_customers_excel():
    """Create a new Excel file with customer data"""
    excel_file = 'new_customers.xlsx'
    
    # Create customer data
    customer_data = {
        'CustomerID': ['C1001', 'C1002', 'C1003', 'C1004', 'C1005'],
        'Name': ['John Smith', 'Emily Johnson', 'Michael Williams', 'Jessica Brown', 'David Jones'],
        'Age': [32, 28, 45, 36, 29],
        'Gender': ['Male', 'Female', 'Male', 'Female', 'Male'],
        'City': ['New York', 'Los Angeles', 'Chicago', 'Houston', 'Phoenix']
    }
    
    customers_df = pd.DataFrame(customer_data)
    
    # Write to Excel
    customers_df.to_excel(excel_file, sheet_name='customers', index=False)
    print(f"Created new customer file: {excel_file}")
    print(f"Added {len(customers_df)} customer records")

if __name__ == "__main__":
    print("Select an option:")
    print("1. Add data to existing Excel or create new one")
    print("2. Create new customer data Excel file")
    
    choice = input("Enter your choice (1 or 2): ")
    
    if choice == '1':
        add_data_to_excel()
    elif choice == '2':
        create_customers_excel()
    else:
        print("Invalid choice. Please run again and select 1 or 2.") 